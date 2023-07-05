# -*- coding: utf-8 -*-
"""
 @Author: eric.zhong
 @Email: ericzhong2010@qq.com
 @Date: 2023/7/4
 @SoftWare: PyCharm
 @FileName: WriteExcelUtil.py
 @Description：
"""

# import library
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import os, time

class WriteExcelUtil:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    # 边界设置
    def table_format_border(self, ws, start_row, end_row, start_col, end_col):
        if start_row != end_row:
            # 内部边界
            for row in tuple(ws[start_row:end_row]):
                for cell in row[start_col - 1:end_col]:
                    # 边框式样参考
                    # 可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
                    cell.border = self.set_border('hair', 'hair', 'hair', 'hair')
            # 左边界
            for cell in [row[start_col - 1] for row in ws[start_row:end_row]]:
                cell.border = self.set_border(cell.border.top.style, cell.border.bottom.style, 'thin',
                                         cell.border.right.style)
            # 右边界
            for cell in [row[end_col - 1] for row in ws[start_row:end_row]]:
                cell.border = self.set_border(cell.border.top.style, cell.border.bottom.style, cell.border.left.style,
                                         'thin')
            # 上边界
            for cell in ws[start_row][start_col - 1:end_col]:
                cell.border = self.set_border('thin', cell.border.bottom.style, cell.border.left.style,
                                         cell.border.right.style)
            # 下边界
            for cell in ws[end_row][start_col - 1:end_col]:
                cell.border = self.set_border(cell.border.top.style, 'thin', cell.border.left.style,
                                         cell.border.right.style)
        else:
            for cell in ws[start_row][start_col - 1:end_col]:
                # 边框式样参考
                # 可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
                cell.border = self.set_border('thin', 'thin', 'thin', 'thin')
        return ws

    # 定义的边界风格
    def set_border(self, t_border, b_border, l_border, r_border, t_color='000000', b_color='000000', l_color='000000', r_color='000000'):
        border = Border(top=Side(border_style=t_border, color=t_color),
                        bottom=Side(border_style=b_border, color=b_color),
                        left=Side(border_style=l_border, color=l_color),
                        right=Side(border_style=r_border, color=r_color))
        return border

    def write_data_to_excel(self, savepath, jsondata):
        # 写入表头
        header = list(jsondata[0].keys())
        for col_num, col_title in enumerate(header, 1):
            self.sheet.cell(row=1, column=col_num, value=col_title)

        # 写入式样
        header_style = {
            'fill': PatternFill(start_color='2400B0', end_color='2400B0', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True)
        }
        header_width = {
            "A": 6,
            "B": 80,
            "C": 21.88,
            "D": 21.88,
            "E": 18,
            "F": 18,
            "G": 11.88,
            "H": 24.88,
            "I": 7.88
        }
        # 获取最大列数
        max_col = self.sheet.max_column

        # 遍历每个列的表头
        for col_num in range(1, max_col + 1):
            # 获取列字母
            col_letter = chr(ord('A') + col_num - 1)
            # 构建表头单元格地址
            header_cell = self.sheet[col_letter + '1']

            # 设置表头单元格样式
            header_cell.fill = header_style['fill']
            header_cell.font = header_style['font']
            self.sheet.column_dimensions[col_letter].width = header_width[col_letter]
        self.table_format_border(self.sheet, 1, self.sheet.max_row, 1, self.sheet.max_column)


        # 写入数据
        for row_num, item in enumerate(jsondata, 2):
            for col_num, key in enumerate(header, 1):
                value = item.get(key)
                self.sheet.cell(row=row_num, column=col_num, value=value)
        # 写入式样
        # (sheet, 开始行, 结束行, 开始列, 结束列)
        self.table_format_border(self.sheet, 2, self.sheet.max_row, 1, self.sheet.max_column)

        # 调整列宽


        # 保存 Excel 文件
        computerName = os.environ.get('computername')
        self.workbook.save('{}\{}_{}.xlsx'.format(savepath, computerName, time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time()))))

writeexcel = WriteExcelUtil()